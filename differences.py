from openpyxl import load_workbook

wb1 = load_workbook(filename='A.xlsx')
wb2 = load_workbook(filename='B.xlsx')

wb1_sheet = wb1['retail']
wb2_sheet = wb2['retail']

a = []
for col in wb1_sheet.iter_cols(min_col=1, max_col=1, values_only=True):
    for row in col:
        if row:
            a.append(row)

b = []
for col in wb2_sheet.iter_cols(min_col=1, max_col=1, values_only=True):
    for row in col:
        if row:
            b.append(row)

if a == b:
    print("The same!!!!")

else:
    if set(b).difference(a):
        for diff in set(b).difference(a):
            indices = [index for index, value in enumerate(b, start=1) if value == diff]
            print(f"{diff} in Row {indices} in Excel B")

    if set(a).difference(b):
        for diff in set(a).difference(b):
            indices = [index for index, value in enumerate(a, start=1) if value == diff]
            print(f"{diff} in Row {indices} in Excel A")
